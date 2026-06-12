#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""주간 스냅샷 행 자동 추가 (2026 시트).
Usage: python3 add_week.py "<YSLim_Investing Summary.xlsx>" [--week W2625] [--test]
- 직전 금요일 종가(stooq.com)와 전주 대비 변동률을 마지막 W블록 2행 아래에 추가.
- 스크립트가 만든 동일 주차 행이 있으면 최신 값으로 덮어씀.
- 색상은 파일에 설정된 조건부 서식이 자동 적용(상승 빨강/하락 파랑).
"""
import sys, re, zipfile, shutil, datetime as dt, urllib.request, urllib.parse, io, csv, html

XLSX = sys.argv[1]
ARGS = sys.argv[2:]
TEST = '--test' in ARGS
SHEET = 'xl/worksheets/sheet4.xml'   # 2026 sheet
NAMES = ['Dow','S&P500','나스닥','TSLA','SOXX','러셀2000','BMNR']
SYMB  = ['^dji','^spx','^ndq','tsla.us','soxx.us','^rut','bmnr.us']


def http_get(url, timeout=60):
    req = urllib.request.Request(url, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36',
        'Accept': 'text/csv,application/json,text/plain,*/*'})
    import time
    for attempt in (1, 2):
        try:
            return urllib.request.urlopen(req, timeout=timeout).read().decode()
        except Exception:
            if attempt == 2: raise
            time.sleep(3)

def last_friday(today=None):
    d = today or dt.date.today()
    while d.weekday() != 4: d -= dt.timedelta(days=1)
    return d

YAHOO = {'^dji':'^DJI','^spx':'^GSPC','^ndq':'^IXIC','tsla.us':'TSLA','soxx.us':'SOXX','^rut':'^RUT','bmnr.us':'BMNR'}

def fetch_close(sym, on_or_before):
    # 1차: stooq
    try:
        raw = http_get(f'https://stooq.com/q/d/l/?s={sym}&i=d', timeout=30)
        best = None
        for r in list(csv.reader(io.StringIO(raw)))[1:]:
            try:
                d = dt.date.fromisoformat(r[0]); c = float(r[4])
            except Exception: continue
            if d <= on_or_before: best = c
        if best is not None: return best
        print(f'  stooq empty for {sym} (resp: {raw[:50]!r}), trying yahoo')
    except Exception as e:
        print(f'  stooq fail {sym}: {e}, trying yahoo')
    # 2차: Yahoo Finance chart API
    try:
        import json as _j
        ysym = YAHOO.get(sym, sym)
        raw = http_get(f'https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(ysym)}?range=1mo&interval=1d', timeout=30)
        res = _j.loads(raw)['chart']['result'][0]
        ts = res.get('timestamp') or []
        closes = res['indicators']['quote'][0].get('close') or []
        best = None
        for t, c in zip(ts, closes):
            if c is None: continue
            if dt.datetime.utcfromtimestamp(t).date() <= on_or_before: best = float(c)
        return round(best, 2) if best is not None else None
    except Exception as e:
        print(f'  yahoo fail {sym}: {e}'); return None

def col_letter(idx):  # 1->A
    s=''
    while idx: idx, r = divmod(idx-1, 26); s = chr(65+r)+s
    return s

def main():
    fri = last_friday()
    iso = fri.isocalendar()
    week = f'W{str(fri.year)[2:]}{iso[1]:02d}'
    for a in ARGS:
        if a.startswith('W') and len(a)==5: week=a
    z = zipfile.ZipFile(XLSX)
    sx = z.read(SHEET).decode()
    # 이 스크립트가 만든 행(inlineStr 라벨)이면 → 최신 값으로 덮어쓰기, 아니면 신규 추가.
    # 사용자가 직접 입력한 라벨(shared string)이 있으면 건드리지 않고 종료.
    existing = re.search(
        rf'<row r="(\d+)"[^>]*>(?:(?!</row>).)*?<is><t>{week}</t>(?:(?!</row>).)*?</row>', sx, re.S)
    if not existing and (f'>{week}<' in sx):
        print(f'{week} exists as a manually entered row; not touching it'); return 0
    rownums = [int(m) for m in re.findall(r'<row r="(\d+)"', sx)]
    maxr = max(rownums)
    newr = int(existing.group(1)) if existing else maxr + 2
    # previous closes: 마지막 스냅샷 행(덮어쓸 대상 행은 제외)
    prev = {}
    lastWrow_xml = None
    for rm in re.finditer(r'<row r="(\d+)"[^>]*>(.*?)</row>', sx, re.S):
        if int(rm.group(1)) == newr: continue  # 자기 자신은 기준에서 제외
        body = rm.group(2)
        cells = dict(re.findall(r'<c r="([A-Z]+)\d+"[^>]*>(?:<f>[^<]*</f>)?<v>([^<]*)</v></c>', body))
        if all(k in cells for k in ('D','G','J','M','P','S','V')):
            try:
                prev = {NAMES[i]: float(cells[c]) for i,c in enumerate(['D','G','J','M','P','S','V'])}
                lastWrow_xml = body
            except ValueError: pass
    if not prev: print('previous snapshot row not found'); return 1
    # style ids from last snapshot row
    sty = dict(re.findall(r'<c r="([A-Z]+)\d+" s="(\d+)"', lastWrow_xml or ''))
    print(f'week={week} friday={fri} row={newr} ({"update" if existing else "new"})')
    closes = {}
    for n, s in zip(NAMES, SYMB):
        closes[n] = 123.45 if TEST else fetch_close(s, fri)
        print(f'  {n}: {closes[n]}')
    cells_xml = []
    def cell(colL, val=None, text=None):
        s = f' s="{sty[colL]}"' if colL in sty else ''
        if text is not None:
            return f'<c r="{colL}{newr}"{s} t="inlineStr"><is><t>{html.escape(text)}</t></is></c>'
        if val is None: return ''
        return f'<c r="{colL}{newr}"{s}><v>{val}</v></c>'
    cells_xml.append(cell('A', text=week))
    base = 3  # C
    for i, n in enumerate(NAMES):
        cN, cV, cP = col_letter(base+i*3), col_letter(base+i*3+1), col_letter(base+i*3+2)
        cells_xml.append(cell(cN, text=n))
        if closes[n] is not None:
            cells_xml.append(cell(cV, val=repr(closes[n])))
            if prev.get(n): cells_xml.append(cell(cP, val=repr(closes[n]/prev[n]-1)))
    newrow = f'<row r="{newr}">' + ''.join(cells_xml) + '</row>'
    if existing:
        sx2 = sx.replace(existing.group(0), newrow, 1)
    else:
        sx2 = sx.replace('</sheetData>', newrow + '</sheetData>', 1)
        sx2 = re.sub(r'(<dimension ref="A1:[A-Z]+)\d+"', lambda m: f'{m.group(1)}{max(maxr,newr)}"', sx2, 1)
    tmp = XLSX + '.tmp.xlsx'
    zo = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED)
    for it in z.infolist():
        zo.writestr(it, sx2 if it.filename == SHEET else z.read(it.filename))
    zo.close(); z.close()
    import openpyxl; openpyxl.load_workbook(tmp, read_only=True)  # validate
    shutil.move(tmp, XLSX)
    print('row updated OK' if existing else 'row added OK')
    return 0


# ===================== Yield 시트 자동 갱신 =====================
YIELD_SHEET = 'xl/worksheets/sheet6.xml'
YIELD_TABLE = 'xl/tables/table2.xml'
EPOCH = dt.date(1899, 12, 30)
TRE_COLS = ['1 Mo','2 Mo','3 Mo','6 Mo','1 Yr','2 Yr','3 Yr','5 Yr','7 Yr','10 Yr','20 Yr','30 Yr']  # -> C..N

def fetch_treasury_month(ym):
    """treasury.gov daily yield curve CSV -> {date: {colname: float}}"""
    base = 'https://home.treasury.gov/resource-center/data-chart-center/interest-rates/daily-treasury-rates.csv'
    try:
        raw = http_get(f'{base}/all/{ym}?type=daily_treasury_yield_curve&field_tdr_date_value={ym}&page&_format=csv')
    except Exception:
        yr = ym[:4]  # 월별 주소 실패 시 연간 전체로 폴백
        raw = http_get(f'{base}/{yr}/all?type=daily_treasury_yield_curve&field_tdr_date_value={yr}&page&_format=csv')
    rows = list(csv.reader(io.StringIO(raw)))
    hdr = rows[0]
    # 1.5 Mo / 4 Mo 등 매칭 컬럼 없는 항목 제외, 이름 정규화
    def norm(h):
        h = h.strip().replace('Month','Mo').replace('Year','Yr')
        return h
    idx = {}
    for i, h in enumerate(hdr):
        n = norm(h)
        if n in TRE_COLS: idx[n] = i
    out = {}
    for r in rows[1:]:
        try: d = dt.datetime.strptime(r[0].strip(), '%m/%d/%Y').date()
        except Exception: continue
        vals = {}
        for n, i in idx.items():
            try: vals[n] = float(r[i])
            except Exception: pass
        if vals: out[d] = vals
    return out

def fetch_effr_range(d0, d1):
    """NY Fed API -> {date: rate}"""
    url = (f'https://markets.newyorkfed.org/api/rates/unsecured/effr/search.json'
           f'?startDate={d0.isoformat()}&endDate={d1.isoformat()}')
    try:
        import json as _j
        raw = http_get(url)
        out = {}
        for r in _j.loads(raw).get('refRates', []):
            out[dt.date.fromisoformat(r['effectiveDate'])] = float(r['percentRate'])
        return out
    except Exception as e:
        print(f'  EFFR fetch fail: {e}'); return {}

def update_yield(test=False):
    z = zipfile.ZipFile(XLSX)
    sx = z.read(YIELD_SHEET).decode()
    tx = z.read(YIELD_TABLE).decode()
    wx = z.read('xl/workbook.xml').decode()
    rowsfull = re.findall(r'<row r="\d+"[^>]*>.*?</row>', sx, re.S)
    last = rowsfull[-1]
    lastr = int(re.search(r'<row r="(\d+)"', last).group(1))
    last_serial = int(re.search(r'<c r="A\d+"[^>]*><v>(\d+)</v>', last).group(1))
    last_date = EPOCH + dt.timedelta(days=last_serial)
    last_fed = float(re.search(r'<c r="B\d+"[^>]*><v>([^<]+)</v>', last).group(1))
    # 수식 셀(O~S)은 표 구조 참조라 행 번호 무관 - 그대로 복사
    formulas = dict(re.findall(r'<c r="([O-S])\d+"[^>]*><f>([^<]*)</f>', last))
    colsty = dict(re.findall(r'<c r="([A-S])\d+" s="(\d+)"', last))  # 마지막 행의 열별 스타일
    fri = last_friday()
    need = []
    d = last_date + dt.timedelta(days=1)
    while d <= fri:
        if d.weekday() < 5: need.append(d)
        d += dt.timedelta(days=1)
    if not need:
        print(f'Yield: up to date ({last_date})'); return
    print(f'Yield: adding {need[0]} ~ {need[-1]}')
    if test:
        tre = {d: {c: 4.0 for c in TRE_COLS} for d in need}
        effr = {d: 3.64 for d in need}
    else:
        months = sorted({f'{d.year}{d.month:02d}' for d in need})
        tre = {}
        for ym in months:
            try: tre.update(fetch_treasury_month(ym))
            except Exception as e: print(f'  treasury {ym} fail: {e}')
        effr = fetch_effr_range(need[0], need[-1])
    newrows, r = [], lastr
    for d in need:
        if d not in tre: continue  # 휴장일
        r += 1
        serial = (d - EPOCH).days
        def sa(c): return f' s="{colsty[c]}"' if c in colsty else ''
        cells = [f'<c r="A{r}"{sa("A")}><v>{serial}</v></c>',
                 f'<c r="B{r}"{sa("B")}><v>{effr.get(d, last_fed)}</v></c>']
        for j, cn in enumerate(TRE_COLS):
            colL = chr(ord('C') + j)
            sattr = sa(colL)
            v = tre[d].get(cn)
            if v is not None: cells.append(f'<c r="{colL}{r}"{sattr}><v>{v}</v></c>')
        for colL in 'OPQRS':
            if colL in formulas:
                cells.append(f'<c r="{colL}{r}"{sa(colL)}><f>{formulas[colL]}</f></c>')
        newrows.append(f'<row r="{r}" spans="1:19">' + ''.join(cells) + '</row>')
    if not newrows:
        print('Yield: no trading-day data fetched'); return
    sx2 = sx.replace('</sheetData>', ''.join(newrows) + '</sheetData>', 1)
    sx2 = re.sub(r'<dimension ref="A1:S\d+"', f'<dimension ref="A1:S{r}"', sx2, count=1)
    tx2 = tx.replace(f'ref="A1:S{lastr}"', f'ref="A1:S{r}"')
    wx2 = wx if 'fullCalcOnLoad' in wx else wx.replace('<calcPr calcId="191029"/>', '<calcPr calcId="191029" fullCalcOnLoad="1"/>')
    tmp = XLSX + '.tmp.xlsx'
    zo = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED)
    for it in z.infolist():
        if it.filename == YIELD_SHEET: zo.writestr(it, sx2)
        elif it.filename == YIELD_TABLE: zo.writestr(it, tx2)
        elif it.filename == 'xl/workbook.xml': zo.writestr(it, wx2)
        else: zo.writestr(it, z.read(it.filename))
    zo.close(); z.close()
    import openpyxl; openpyxl.load_workbook(tmp, read_only=True)
    shutil.move(tmp, XLSX)
    print(f'Yield: {len(newrows)} rows added (through {need[-1]})')



# ===================== 서식/수식 보정 (멱등) =====================
RED_DXF = '<dxf><font><color rgb="FFFF0000"/></font></dxf>'
BLUE_DXF = '<dxf><font><color rgb="FF0000FF"/></font></dxf>'
TRIPLETS = [('C','E'),('F','H'),('I','K'),('L','N'),('O','Q'),('R','T'),('U','W')]

def _rewrite(zin, replacements):
    tmp = XLSX + '.tmp.xlsx'
    zo = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        zo.writestr(it, replacements.get(it.filename, zin.read(it.filename)))
    zo.close(); zin.close()
    import openpyxl; openpyxl.load_workbook(tmp, read_only=True)
    shutil.move(tmp, XLSX)

def ensure_set_colors():
    """2026 시트: 주차 스냅샷의 이름+값+% 3칸을 %부호에 따라 세트로 빨강/파랑."""
    z = zipfile.ZipFile(XLSX)
    sx = z.read(SHEET).decode()
    if 'priority="9101"' in sx:
        z.close(); print('set-colors: already applied'); return
    styles = z.read('xl/styles.xml').decode()
    m = re.search(r'<dxfs count="(\d+)">(.*?)</dxfs>', styles, re.S)
    dxfs = re.findall(r'<dxf>.*?</dxf>', m.group(2), re.S) if m else []
    def get_id(d):
        return dxfs.index(d) if d in dxfs else None
    red, blue = get_id(RED_DXF), get_id(BLUE_DXF)
    if red is None or blue is None:
        add = ('' if red is not None else RED_DXF) + ('' if blue is not None else BLUE_DXF)
        if m:
            styles = styles.replace(f'<dxfs count="{m.group(1)}">', f'<dxfs count="{int(m.group(1)) + add.count("<dxf>")}">', 1)
            styles = styles.replace('</dxfs>', add + '</dxfs>', 1)
        else:
            block = f'<dxfs count="{add.count("<dxf>")}">{add}</dxfs>'
            styles = styles.replace('<tableStyles', block + '<tableStyles', 1) if '<tableStyles' in styles else styles.replace('</styleSheet>', block + '</styleSheet>', 1)
        if red is None: red = len(dxfs); dxfs.append(RED_DXF)
        if blue is None: blue = len(dxfs); dxfs.append(BLUE_DXF)
    blocks, pr = [], 9101
    for c0, cp in TRIPLETS:
        sq = f'{c0}1:{cp}3000'
        blocks.append(
            f'<conditionalFormatting sqref="{sq}">'
            f'<cfRule type="expression" dxfId="{red}" priority="{pr}"><formula>AND(ISNUMBER(${cp}1),${cp}1&gt;0,${cp}1&lt;1)</formula></cfRule>'
            f'<cfRule type="expression" dxfId="{blue}" priority="{pr+1}"><formula>AND(ISNUMBER(${cp}1),${cp}1&lt;0,${cp}1&gt;-1)</formula></cfRule>'
            f'</conditionalFormatting>')
        pr += 2
    cf = ''.join(blocks)
    for anchor in ['<hyperlinks', '<printOptions', '<pageMargins', '<pageSetup', '<drawing']:
        if anchor in sx:
            sx = sx.replace(anchor, cf + anchor, 1); break
    _rewrite(z, {SHEET: sx, 'xl/styles.xml': styles})
    print('set-colors: applied')

def ensure_yield_formulas():
    """Yield 시트: O~S 수식이 빠진 데이터 행에 수식 채워넣기."""
    z = zipfile.ZipFile(XLSX)
    sx = z.read(YIELD_SHEET).decode()
    donor = None
    for rm in re.finditer(r'<row r="\d+"[^>]*>.*?</row>', sx, re.S):
        if re.search(r'<c r="O\d+"[^>]*><f>', rm.group(0)): donor = rm.group(0)
    if donor is None:
        z.close(); print('yield-formulas: no donor row'); return
    formulas = dict(re.findall(r'<c r="([O-S])\d+"[^>]*><f>([^<]*)</f>', donor))
    dsty = dict(re.findall(r'<c r="([O-S])\d+" s="(\d+)"', donor))
    fixed = 0
    def fix(rm):
        nonlocal fixed
        row = rm.group(0)
        rno = rm.group(1)
        if re.search(r'<c r="O\d+"', row): return row          # 이미 있음
        if not re.search(r'<c r="B\d+"[^>]*><v>', row): return row  # 데이터 행 아님
        if not re.search(r'<c r="A\d+"[^>]*><v>\d+</v>', row): return row
        cells = ''
        for colL in 'OPQRS':
            if colL in formulas:
                s = f' s="{dsty[colL]}"' if colL in dsty else ''
                cells += f'<c r="{colL}{rno}"{s}><f>{formulas[colL]}</f></c>'
        fixed += 1
        return row[:-len('</row>')] + cells + '</row>'
    sx2 = re.sub(r'<row r="(\d+)"[^>]*>.*?</row>', fix, sx, flags=re.S)
    if not fixed:
        z.close(); print('yield-formulas: all rows OK'); return
    wx = z.read('xl/workbook.xml').decode()
    wx2 = wx if 'fullCalcOnLoad' in wx else wx.replace('<calcPr calcId="191029"/>', '<calcPr calcId="191029" fullCalcOnLoad="1"/>')
    _rewrite(z, {YIELD_SHEET: sx2, 'xl/workbook.xml': wx2})
    print(f'yield-formulas: fixed {fixed} rows')

def run_all():
    rc = main()
    try:
        update_yield(test=TEST)
    except Exception as e:
        print(f'Yield update failed: {e}')
    for fn in (ensure_yield_formulas, ensure_set_colors):
        try: fn()
        except Exception as e: print(f'{fn.__name__} failed: {e}')
    return rc

sys.exit(run_all())
