#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""YSLim Investing Summary - static site builder (v2, single workbook).
Usage: python3 build_site.py "<YSLim_Investing Summary.xlsx>" <out_dir>
- 시트: Orientation / Summary_~2024 / 2025 / 2026 / Factset / Yield
- 흰색 폰트 셀(비공개 메모)은 사이트에서 제외된다.
"""
import sys, os, re, zipfile, datetime, json, shutil, html as H

XLSX, SITE = sys.argv[1], sys.argv[2]
TOOLS = os.path.dirname(os.path.abspath(__file__))
IMGDIR = os.path.join(SITE, 'assets', 'img')
os.makedirs(IMGDIR, exist_ok=True)
shutil.copy(os.path.join(TOOLS, 'chart.umd.js'), os.path.join(SITE, 'assets', 'chart.umd.js'))


# ---- Q&A (giscus = GitHub Discussions 댓글) 설정 ----
# giscus.app에서 저장소 입력 후 발급되는 두 값을 붙여넣으면 Q&A 탭이 활성화됨.
GISCUS = {
    'repo': 'YSLim33/YSLim-Investing-Summary',
    'repo_id': 'R_kgDOS4J4WQ',
    'category': 'Q&A',
    'category_id': 'DIC_kwDOS4J4Wc4C_DHv',
}
CONTACT_EMAIL = 'june.lim33@gmail.com'

LINKS6 = [
 ("Financial condition & BEI & Sales/Inventory", "https://fredaccount.stlouisfed.org/public/dashboard/111622", "FRED 대시보드 111622"),
 ("Yield & Bond", "https://fredaccount.stlouisfed.org/public/dashboard/111624", "FRED 대시보드 111624"),
 ("FED Balance Sheet", "https://fredaccount.stlouisfed.org/public/dashboard/111626", "FRED 대시보드 111626"),
 ("환율 (FX)", "https://fredaccount.stlouisfed.org/public/dashboard/134732", "FRED 대시보드 134732"),
 ("GDPNow", "https://www.atlantafed.org/cqer/research/gdpnow#Tab3", "Atlanta Fed"),
 ("CME FedWatch", "https://www.cmegroup.com/markets/interest-rates/cme-fedwatch", "CME Group"),
]

CSS = """
:root{--bg:#0f1419;--card:#1a2129;--border:#2a3441;--text:#e6edf3;--muted:#8b98a5;--accent:#4da3ff;}
*{box-sizing:border-box;}
body{background:var(--bg);color:var(--text);font-family:'Segoe UI','Malgun Gothic',system-ui,sans-serif;margin:0;padding:0 16px 60px;}
header{display:flex;flex-wrap:wrap;align-items:baseline;gap:18px;padding:18px 4px;border-bottom:1px solid var(--border);margin-bottom:20px;}
header h1{font-size:19px;margin:0;}
nav{display:flex;flex-wrap:wrap;gap:4px;}
nav a{color:var(--muted);text-decoration:none;font-size:14px;padding:4px 10px;border-radius:6px;}
nav a:hover{color:var(--text);background:var(--card);}
nav a.on{color:var(--accent);font-weight:600;}
.wrap{max-width:1060px;margin:0 auto;}
.card{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px 18px;margin-bottom:16px;}
.card h2{font-size:16px;color:var(--accent);margin:0 0 10px;}
.week{border-left:3px solid var(--accent);}
table.snap{border-collapse:collapse;width:100%;margin:6px 0 12px;font-size:13px;}
table.snap th,table.snap td{border:1px solid var(--border);padding:5px 9px;text-align:right;}
table.snap th{color:var(--muted);font-weight:600;}
.up{color:#f85149;}.dn{color:#4da3ff;}
.line{font-size:13.5px;line-height:1.65;color:var(--text);margin:2px 0;white-space:pre-wrap;}
.line a{color:var(--accent);}
img.emb{max-width:100%;border-radius:8px;border:1px solid var(--border);margin:8px 6px 8px 0;vertical-align:top;}
.imgrow{display:flex;flex-wrap:wrap;gap:8px;}
.linkgrid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:12px;}
.linkgrid a{display:block;background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px 16px;text-decoration:none;}
.linkgrid a:hover{border-color:var(--accent);}
.linkgrid .t{color:var(--text);font-weight:600;font-size:14.5px;}
.linkgrid .s{color:var(--muted);font-size:12.5px;margin-top:4px;}
.muted{color:var(--muted);font-size:12.5px;}
details{margin-bottom:14px;}
details summary{cursor:pointer;font-size:15px;color:var(--accent);padding:10px 14px;background:var(--card);border:1px solid var(--border);border-radius:10px;}
details[open] summary{border-radius:10px 10px 0 0;}
details .body{border:1px solid var(--border);border-top:none;border-radius:0 0 10px 10px;padding:12px 16px;background:#141a21;}
@keyframes flashhl{0%{background:rgba(77,163,255,.5);}100%{background:transparent;}}
details.flash summary{animation:flashhl 1.9s ease-out;}
"""
# 상승=빨강(.up), 하락=파랑(.dn) — 한국식 컬러

NAV_ITEMS = []  # 워크북 로드 후 연도 시트를 발견해 채움

def page(title, active, body):
    nav = ''.join(f'<a href="{h}" class="{"on" if k==active else ""}">{t}</a>' for t,h,k in NAV_ITEMS)
    return f"""<!DOCTYPE html><html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{H.escape(title)} · YSLim Investing</title><style>{CSS}</style></head>
<body><div class="wrap"><header><h1>YSLim Investing Summary</h1><nav>{nav}</nav></header>
{body}
<p class="muted">생성: {datetime.date.today().isoformat()} · 본 자료는 개인 기록·공유용이며 투자 권유가 아닙니다. 투자 판단과 책임은 각자에게 있습니다.</p>
</div></body></html>"""

def esc(x): return H.escape(str(x))

# ---- 검색(Search) 인덱스 ----
SEARCH_RECORDS = []
def slugify(s):
    s = re.sub(r'[^0-9A-Za-z가-힣]+', '-', str(s)).strip('-')
    return s or 'sec'
def html_to_text(s):
    s = re.sub(r'<br\s*/?>', ' ', s)
    s = re.sub(r'<[^>]+>', ' ', s)
    return re.sub(r'\s+', ' ', H.unescape(s)).strip()
def add_record(title, url, group, text, external=False, week=None):
    text = (text or '').strip()
    if not text: return
    rec = {'t': title, 'u': url, 'g': group, 'x': text}
    if external: rec['e'] = 1
    if week is not None: rec['w'] = week
    SEARCH_RECORDS.append(rec)
def index_blocks(blocks, page_url, group):
    for lbl, bh in blocks:
        m = re.match(r'W(\d+)', str(lbl))
        add_record(str(lbl), f'{page_url}#{slugify(lbl)}', group,
                   html_to_text(bh), week=int(m.group(1)) if m else None)

def sheet_xml_map(zf):
    wbx = zf.read('xl/workbook.xml').decode()
    sheets = re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wbx)
    rels = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', zf.read('xl/_rels/workbook.xml.rels').decode()))
    return {name: rels[rid].split('/')[-1] for name, rid in sheets}

def sheet_assets(zf, sheetfile, prefix):
    imgs, links = [], {}
    relp = f'xl/worksheets/_rels/{sheetfile}.rels'
    if relp not in zf.namelist(): return imgs, links
    rels = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', zf.read(relp).decode()))
    sx = zf.read(f'xl/worksheets/{sheetfile}').decode()
    for ref, rid in re.findall(r'<hyperlink ref="([A-Z]+\d+)" r:id="(rId\d+)"', sx):
        if rid in rels: links[int(re.search(r'\d+', ref).group())] = rels[rid]
    drawing = [v for v in rels.values() if 'drawing' in v]
    if drawing:
        dname = drawing[0].split('/')[-1]
        dr = zf.read(f'xl/drawings/{dname}').decode()
        drels = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', zf.read(f'xl/drawings/_rels/{dname}.rels').decode()))
        for a in re.findall(r'<xdr:(?:two|one)CellAnchor.*?</xdr:(?:two|one)CellAnchor>', dr, re.S):
            rowm = re.search(r'<xdr:from>.*?<xdr:row>(\d+)</xdr:row>', a, re.S)
            em = re.search(r'r:embed="(rId\d+)"', a)
            if not (rowm and em and em.group(1) in drels): continue
            mediafile = drels[em.group(1)].split('/')[-1]
            out = f'{prefix}_{mediafile}'
            outp = os.path.join(IMGDIR, out)
            if not os.path.exists(outp):
                with open(outp, 'wb') as f: f.write(zf.read(f'xl/media/{mediafile}'))
            imgs.append((int(rowm.group(1)) + 1, 0, f'assets/img/{out}'))
    return imgs, links

def is_white(cell):
    col = cell.font.color
    if col is None: return False
    t = getattr(col, 'type', None)
    if t == 'rgb' and str(col.rgb).upper() in ('FFFFFFFF', '00FFFFFF'): return True
    if t == 'theme' and col.theme == 0 and not getattr(col, 'tint', 0): return True
    return False

ASSETS = ['Dow','S&P500','나스닥','TSLA','SOXX','러셀2000','BMNR']
def fmt_pct(v):
    try: v = float(v)
    except: return ''
    cls = 'up' if v >= 0 else 'dn'
    return f'<span class="{cls}">{v*100:+.2f}%</span>'

def render_rows(rows, imgs, links, r0, r1, week_mode=True):
    out = []
    imap = {}
    for r, c, p in imgs:
        if r0 <= r <= r1: imap.setdefault(r, []).append(p)
    i = r0
    while i <= r1:
        row = rows[i-1] if i-1 < len(rows) else ()
        cells = [c for c in row if c is not None and str(c).strip() != '']
        first = str(row[0]).strip() if row and row[0] is not None else ''
        if week_mode and re.match(r'^W\d{4}', first):
            head, vals = [], []
            for j, name in enumerate(ASSETS):
                base = 2 + j*3
                if base+1 < len(row) and isinstance(row[base+1], (int, float)):
                    head.append(f'<th>{esc(row[base] or name)}</th>')
                    pct = fmt_pct(row[base+2]) if base+2 < len(row) else ''
                    v = row[base+1]
                    vs = f'{v:,.2f}' if isinstance(v,(int,float)) else esc(v)
                    vals.append(f'<td>{vs}<br>{pct}</td>')
            out.append(f'<h2>{esc(first)}</h2><table class="snap"><tr>{"".join(head)}</tr><tr>{"".join(vals)}</tr></table>')
        elif cells:
            txt = '  '.join(esc(c) for c in cells if not isinstance(c, datetime.datetime))
            if i in links:
                u = esc(links[i]); txt += f' <a href="{u}" target="_blank">[link]</a>'
            if txt.strip(): out.append(f'<div class="line">{txt}</div>')
        if i in imap:
            out.append('<div class="imgrow">' + ''.join(f'<a href="{p}" target="_blank"><img class="emb" src="{p}" loading="lazy" style="max-width:480px"></a>' for p in imap[i]) + '</div>')
        i += 1
    return '\n'.join(out)

def render_summary_sheet(rows, imgs, links):
    starts = []
    for i, row in enumerate(rows, 1):
        first = str(row[0]).strip() if row and row[0] is not None else ''
        if re.match(r'^W\d{4}', first): starts.append(i)
    if not starts: return [('전체', render_rows(rows, imgs, links, 1, len(rows), week_mode=False))]
    blocks, seen = [], {}
    for k, s in enumerate(starts):
        e = (starts[k+1]-1) if k+1 < len(starts) else len(rows)
        lbl = str(rows[s-1][0]).strip()
        if lbl in seen: seen[lbl] += 1; lbl = f'{lbl}({seen[lbl]})'
        else: seen[lbl] = 1
        blocks.append((lbl, render_rows(rows, imgs, links, s, e)))
    return blocks

# 검색 결과 등에서 #해시로 들어오면 해당 주차를 펼치고 스크롤·하이라이트
ARCHIVE_HASH_JS = """<script>
(function(){function go(){var h=decodeURIComponent((location.hash||'').slice(1));if(!h)return;var el=document.getElementById(h);if(!el)return;if(el.tagName==='DETAILS')el.open=true;el.scrollIntoView({behavior:'smooth',block:'start'});el.classList.remove('flash');void el.offsetWidth;el.classList.add('flash');}
window.addEventListener('DOMContentLoaded',go);window.addEventListener('hashchange',go);})();
</script>"""

def archive_page(blocks, latest_open=False):
    parts = []
    for k, (lbl, bh) in enumerate(reversed(blocks)):
        op = ' open' if (latest_open and k == 0) else ''
        parts.append(f'<details id="{slugify(lbl)}"{op}><summary>{esc(lbl)}</summary><div class="body week">{bh}</div></details>')
    return (f'<p class="muted">{len(blocks)}개 주차 · 최신순 · 제목을 누르면 펼쳐집니다</p>'
            + '\n'.join(parts) + ARCHIVE_HASH_JS)

# ---------- load ----------
import openpyxl
print('loading workbook (styles for privacy filter)...')
wb = openpyxl.load_workbook(XLSX, data_only=True)
zf = zipfile.ZipFile(XLSX)
smap = sheet_xml_map(zf)

def masked_rows(sheet):
    ws = wb[sheet]; out = []
    for row in ws.iter_rows():
        out.append(tuple(None if (c.value is not None and is_white(c)) else c.value for c in row))
    return out

# 연도 시트 자동 발견 (2025, 2026, 2027, ...)
years = sorted([n for n in wb.sheetnames if re.fullmatch(r'20\d{2}', n)], reverse=True)
NAV_ITEMS.extend([('최신 브리핑','index.html','index'),('퀵 링크','links.html','links'),
    ('Orientation','orientation.html','orient')])
NAV_ITEMS.extend([(y, f'archive-{y}.html', y) for y in years])
NAV_ITEMS.extend([('2023–2024','archive-2024.html','2024'),
    ('Factset','factset.html','factset'),('Yield','yield.html','yield'),
    ('Q&A','qna.html','qna'),('🔍 검색','search.html','search')])

year_blocks = {}
for y in years:
    rws = masked_rows(y); im, lk = sheet_assets(zf, smap[y], 'y' + y[2:])
    bl = render_summary_sheet(rws, im, lk)
    year_blocks[y] = bl
    index_blocks(bl, f'archive-{y}.html', f'{y} 주간')
    open(os.path.join(SITE, f'archive-{y}.html'), 'w', encoding='utf-8').write(
        page(f'{y} Archive', y, archive_page(bl)))
    print(y + ':', len(bl))
rows24 = masked_rows('Summary_~2024'); img24, lnk24 = sheet_assets(zf, smap['Summary_~2024'], 'y24')
b24 = render_summary_sheet(rows24, img24, lnk24)
index_blocks(b24, 'archive-2024.html', '2023–2024 주간')
open(os.path.join(SITE,'archive-2024.html'),'w',encoding='utf-8').write(page('2023–2024 Archive','2024', archive_page(b24)))
print('~2024:', len(b24))
# 최신 연도에 블록이 없으면(연초) 직전 연도 사용
latest_year = next(y for y in years if year_blocks[y])
b26 = year_blocks[latest_year]
# Orientation
rowsor = masked_rows('Orientation'); imgor, lnkor = sheet_assets(zf, smap['Orientation'], 'ori')
orient_html = '<div class="card">' + render_rows(rowsor, imgor, lnkor, 1, len(rowsor), week_mode=False) + '</div>'
for ln in re.findall(r'<div class="line">(.*?)</div>', orient_html, re.S):
    t = html_to_text(ln)
    if len(t) >= 4:
        add_record(t[:42] + ('…' if len(t) > 42 else ''), 'orientation.html', 'Orientation', t)
open(os.path.join(SITE,'orientation.html'),'w',encoding='utf-8').write(page('Orientation','orient',
  '<p class="muted">지표 체계 설명서 — Orientation 시트</p>' + orient_html))
# links
for t,u,s in LINKS6:
    add_record(t, u, '퀵 링크', f'{t} {s}', external=True)
lg = ''.join(f'<a href="{u}" target="_blank"><div class="t">{i+1}. {esc(t)}</div><div class="s">{esc(s)}</div></a>' for i,(t,u,s) in enumerate(LINKS6))
open(os.path.join(SITE,'links.html'),'w',encoding='utf-8').write(page('퀵 링크','links',
  f'<p class="muted">Summary 시트 상단 고정 링크 6개</p><div class="linkgrid">{lg}</div>'))
# Q&A 페이지 (giscus)
if GISCUS['repo_id'] and GISCUS['category_id']:
    giscus_embed = f"""<script src="https://giscus.app/client.js"
  data-repo="{GISCUS['repo']}" data-repo-id="{GISCUS['repo_id']}"
  data-category="{GISCUS['category']}" data-category-id="{GISCUS['category_id']}"
  data-mapping="specific" data-term="qna" data-strict="0"
  data-reactions-enabled="1" data-input-position="top"
  data-theme="dark" data-lang="ko" crossorigin="anonymous" async></script>"""
else:
    giscus_embed = '<div class="card"><p class="line">Q&amp;A 게시판 준비 중입니다. (giscus 설정 대기)</p></div>'
qna_body = f"""
<div class="card"><h2>문의 및 토론</h2>
<p class="line">Summary 내용에 대한 질문이나 의견을 남겨주세요. 답변이 달리면 이 페이지에서 확인할 수 있습니다.</p>
<p class="line muted">· 글을 쓰려면 GitHub 계정 로그인이 필요합니다 (무료 가입). 로그인 없이 읽기는 가능합니다.<br>
· GitHub 계정이 없으면 이메일로 보내주세요: <a href="mailto:{CONTACT_EMAIL}?subject=[Investing Summary 문의]" style="color:var(--accent)">{CONTACT_EMAIL}</a></p>
</div>
{giscus_embed}
"""
open(os.path.join(SITE,'qna.html'),'w',encoding='utf-8').write(page('Q&A','qna', qna_body))

# factset / yield
fs_rows = [tuple(c.value for c in r) for r in wb['Factset'].iter_rows()]
fs = [[r[0].strftime('%Y-%m-%d')] + [float(x) if isinstance(x,(int,float)) else None for x in r[1:8]] for r in fs_rows if r and hasattr(r[0],'strftime')]
add_record('Factset — 실적·가이던스·FWD PER', 'factset.html', 'Factset',
    'Factset earnings growth positive negative guidance FWD 12M PER forward 5Y 10Y average 실적 성장 긍정 부정 가이던스 밸류에이션 멀티플 ' + ' '.join(r[0] for r in fs))
yd_rows = [tuple(c.value for c in r) for r in wb['Yield'].iter_rows()]
yh = [str(x) for x in yd_rows[0][:19]]
yd = []
for r in yd_rows[1:]:
    if r[0] is None or not hasattr(r[0],'strftime'): continue
    yd.append([r[0].strftime('%Y-%m-%d')] + [float(x) if isinstance(x,(int,float)) else None for x in r[1:19]])
add_record('Yield — 미국채 금리·수익률 곡선', 'yield.html', 'Yield',
    'Yield 금리 국채 수익률 곡선 yield curve treasury spread 장단기 EFFR SOFR 10Y2Y 10Y3M ' + ' '.join(yh) + ' ' + ' '.join(r[0] for r in yd))
CHART_HDR = '<script src="assets/chart.umd.js"></script>'
JSBASE = "Chart.defaults.color='#8b98a5';Chart.defaults.borderColor='#2a3441';"
fs_table_rows = ''.join('<tr>' + ''.join(f'<td>{("" if v is None else (f"{v:.1f}" if isinstance(v,float) and i!=4 else (f"{v*100:.1f}%" if i==4 and v is not None else v)))}</td>' for i,v in enumerate(r)) + '</tr>' for r in reversed(fs[-30:]))
fs_body = f"""{CHART_HDR}
<div class="card"><h2>Earnings Growth & Positive Guidance %</h2><div style="height:330px;position:relative"><canvas id="c1"></canvas></div></div>
<div class="card"><h2>FWD 12M PER vs 5Y/10Y 평균</h2><div style="height:330px;position:relative"><canvas id="c2"></canvas></div></div>
<div class="card"><h2>최근 30개 리포트</h2><div style="overflow-x:auto"><table class="snap"><tr><th>Date</th><th>Earn growth%</th><th>Neg G</th><th>Pos G</th><th>Pos %</th><th>FWD PER</th><th>5Y avg</th><th>10Y avg</th></tr>{fs_table_rows}</table></div></div>
<script>const F={json.dumps(fs)};{JSBASE}
const L=F.map(r=>r[0]);
new Chart(c1,{{data:{{labels:L,datasets:[{{type:'bar',label:'Earnings growth %',data:F.map(r=>r[1]),backgroundColor:F.map(r=>r[1]>=0?'rgba(248,81,73,.55)':'rgba(77,163,255,.55)')}},{{type:'line',label:'Positive guidance %(우)',data:F.map(r=>r[4]==null?null:r[4]*100),borderColor:'#d29922',pointRadius:0,yAxisID:'y2'}}]}},options:{{maintainAspectRatio:false,scales:{{x:{{ticks:{{maxTicksLimit:10}}}},y2:{{position:'right',min:0,max:100,grid:{{drawOnChartArea:false}}}}}}}}}});
new Chart(c2,{{type:'line',data:{{labels:L,datasets:[{{label:'FWD 12M PER',data:F.map(r=>r[5]),borderColor:'#4da3ff',pointRadius:0}},{{label:'5Y avg',data:F.map(r=>r[6]),borderColor:'#3fb950',pointRadius:0,borderDash:[5,4]}},{{label:'10Y avg',data:F.map(r=>r[7]),borderColor:'#d29922',pointRadius:0,borderDash:[2,3]}}]}},options:{{maintainAspectRatio:false,scales:{{x:{{ticks:{{maxTicksLimit:10}}}}}}}}}});
</script>"""
open(os.path.join(SITE,'factset.html'),'w',encoding='utf-8').write(page('Factset','factset', fs_body))
yd_last = yd[-1]
# Yield_Chart 시트의 두 차트를 Yield 데이터로 재현 (Today/월평균, Today~Today-4)
full = [r for r in yd if all(v is not None for v in r[1:14])]
mats13 = yh[1:14]
def curve(r): return r[1:14]
def avgw(a, b):
    seg = [curve(r) for r in (full[a:b] if b != 0 else full[a:])]
    return [round(sum(c)/len(c), 4) for c in zip(*seg)]
M = 21
avg_sets = [('1개월 평균', avgw(-M, 0)), ('1-2개월 평균', avgw(-2*M, -M)), ('2-3개월 평균', avgw(-3*M, -2*M)),
            ('3-4개월 평균', avgw(-4*M, -3*M)), ('4-5개월 평균', avgw(-5*M, -4*M)), ('5-6개월 평균', avgw(-6*M, -5*M))]
days5 = [(('Today' if i == 0 else f'Today -{i}') + f' ({full[-1-i][0]})', curve(full[-1-i])) for i in range(5)]
AVG_COLORS = ['#f4a261','#e9c46a','#8ab17d','#6a994e','#577590','#9d8189']
DAY_COLORS = ['#d62828','#4da3ff','#8b98a5','#577590','#3a4a5c']
def ds(label, data, color, width=1.3):
    return {'label': label, 'data': data, 'borderColor': color, 'backgroundColor': color,
            'borderWidth': width, 'pointRadius': 2, 'tension': 0.3}
c1_ds = [ds(f'Today ({full[-1][0]})', curve(full[-1]), '#d62828', 2.2)] + [ds(n, v, c) for (n, v), c in zip(avg_sets, AVG_COLORS)]
c2_ds = [ds(n, v, DAY_COLORS[i], 2.2 if i == 0 else 1.3) for i, (n, v) in enumerate(days5)]
yc_head = ''.join(f'<th>{esc(h)}</th>' for h in yh[:14])
yc_vals = ''.join(f'<td>{"" if v is None else f"{v:.2f}"}</td>' for v in yd_last[1:14])
yd_body = f"""{CHART_HDR}
<div class="card"><h2>Today 기준 평균(월)</h2><div style="height:360px;position:relative"><canvas id="c1"></canvas></div>
<div class="muted">Yield_Chart 시트 차트 1과 동일 — Today + 1~6개월 구간 평균 곡선</div></div>
<div class="card"><h2>Today ~ Today-4</h2><div style="height:360px;position:relative"><canvas id="c2"></canvas></div>
<div class="muted">Yield_Chart 시트 차트 2와 동일 — 최근 5영업일 곡선</div></div>
<div class="card"><h2>금리 추이 (10Y · 2Y · 30Y · 10Y2Y)</h2><div style="height:340px;position:relative"><canvas id="c3"></canvas></div></div>
<div class="card"><h2>최신 수치 ({yd_last[0]})</h2>
<div style="overflow-x:auto"><table class="snap"><tr><th></th>{yc_head}</tr><tr><td>{yd_last[0]}</td>{yc_vals}</tr></table></div></div>
<script>const Y={json.dumps(yd)};{JSBASE}
const mats={json.dumps(mats13)};
new Chart(c1,{{type:'line',data:{{labels:mats,datasets:{json.dumps(c1_ds, ensure_ascii=False)}}},options:{{maintainAspectRatio:false,interaction:{{mode:'index',intersect:false}}}}}});
new Chart(c2,{{type:'line',data:{{labels:mats,datasets:{json.dumps(c2_ds, ensure_ascii=False)}}},options:{{maintainAspectRatio:false,interaction:{{mode:'index',intersect:false}}}}}});
const L=Y.map(r=>r[0]);
new Chart(c3,{{type:'line',data:{{labels:L,datasets:[{{label:'10Y',data:Y.map(r=>r[11]),borderColor:'#4da3ff',pointRadius:0,borderWidth:1.6}},{{label:'2Y',data:Y.map(r=>r[7]),borderColor:'#3fb950',pointRadius:0,borderWidth:1.3}},{{label:'30Y',data:Y.map(r=>r[13]),borderColor:'#bc8cff',pointRadius:0,borderWidth:1.3}},{{label:'10Y2Y(우)',data:Y.map(r=>r[14]),borderColor:'#d29922',pointRadius:0,borderWidth:1.3,yAxisID:'y2'}}]}},options:{{maintainAspectRatio:false,interaction:{{mode:'index',intersect:false}},scales:{{x:{{ticks:{{maxTicksLimit:10}}}},y2:{{position:'right',grid:{{drawOnChartArea:false}}}}}}}}}});
</script>"""
open(os.path.join(SITE,'yield.html'),'w',encoding='utf-8').write(page('Yield','yield', yd_body))
# index
latest_lbl, latest_html = b26[-1]
lg_small = ''.join(f'<a href="{u}" target="_blank"><div class="t">{esc(t)}</div><div class="s">{esc(s)}</div></a>' for t,u,s in LINKS6)
idx = f"""<div class="card week"><h2 style="font-size:18px">최신 주간 브리핑 — {esc(latest_lbl)}</h2>{latest_html}</div>
<div class="card"><h2>퀵 링크</h2><div class="linkgrid">{lg_small}</div></div>
<p class="muted">과거 주차: {' · '.join(f'<a href="archive-{y}.html" style="color:var(--accent)">{y}</a>' for y in years)} · <a href="archive-2024.html" style="color:var(--accent)">2023–2024</a></p>"""
open(os.path.join(SITE,'index.html'),'w',encoding='utf-8').write(page('최신 브리핑','index', idx))
# ---- 검색 페이지 (search.html) : 한/영 OR 검색, 인덱스 임베드 ----
SEARCH_CSS = """<style>
.sbox{position:relative;margin:8px 0 14px;}
.sbox input{width:100%;font-size:16px;color:var(--text);background:var(--bg);border:1px solid var(--border);border-radius:12px;padding:14px 42px 14px 16px;outline:none;}
.sbox input:focus{border-color:var(--accent);}
.sbox .clr{position:absolute;right:8px;top:50%;transform:translateY(-50%);width:30px;height:30px;line-height:28px;text-align:center;cursor:pointer;color:var(--muted);font-size:20px;background:none;border:none;display:none;}
.sbox .clr:hover{color:var(--text);}
.hint{color:var(--muted);font-size:12.5px;margin:0 2px 8px;line-height:1.6;}
.hint b{color:var(--text);}
.chips{display:flex;flex-wrap:wrap;gap:6px;margin:8px 0 2px;}
.chip{font-size:12.5px;color:var(--text);background:var(--bg);border:1px solid var(--border);border-radius:999px;padding:5px 11px;cursor:pointer;}
.chip:hover{border-color:var(--accent);color:var(--accent);}
.chip .n{color:var(--muted);margin-left:6px;font-size:11.5px;}
.sumline{color:var(--muted);font-size:13px;margin:16px 2px 10px;}
.sumline b{color:var(--text);}
a.res{display:block;background:var(--card);border:1px solid var(--border);border-radius:10px;padding:11px 14px;margin-bottom:9px;text-decoration:none;}
a.res:hover,a.res.sel{border-color:var(--accent);background:#1f2a35;}
.res .top{display:flex;align-items:center;gap:8px;margin-bottom:5px;flex-wrap:wrap;}
.badge{font-size:11px;color:var(--accent);border:1px solid var(--accent);border-radius:6px;padding:1px 7px;white-space:nowrap;}
.res .ttl{font-weight:600;font-size:14.5px;color:var(--text);}
.res .arrow{margin-left:auto;color:var(--muted);font-size:11.5px;white-space:nowrap;}
.snip{font-size:13px;line-height:1.65;color:var(--muted);}
.snip mark{background:rgba(210,153,34,.32);color:#ffd866;border-radius:3px;padding:0 2px;}
.empty{color:var(--muted);font-size:13.5px;padding:16px 2px;line-height:1.7;}
kbd{background:var(--card);border:1px solid var(--border);border-radius:4px;padding:1px 6px;font-size:11px;color:var(--muted);}
</style>"""

SEARCH_JS = r'''(function(){
var IDX=SEARCH_INDEX;for(var i=0;i<IDX.length;i++)IDX[i]._l=IDX[i].x.toLowerCase();
var qEl=document.getElementById('q'),out=document.getElementById('out'),clr=document.getElementById('clr');
var sel=-1,cur=[];
function esc(s){return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
function rxesc(s){return s.replace(/[.*+?^${}()|[\]\\]/g,'\\$&');}
function terms(q){var a=q.toLowerCase().split(/\s+/).filter(Boolean),u=[];for(var i=0;i<a.length;i++)if(u.indexOf(a[i])<0)u.push(a[i]);return u;}
function run(q){
 var ts=terms(q);cur=[];sel=-1;
 if(!ts.length){render(ts,null);return;}
 var rx=new RegExp('('+ts.map(rxesc).join('|')+')','gi');
 for(var i=0;i<IDX.length;i++){
  var r=IDX[i],L=r._l,matched=[],occ=0;
  for(var j=0;j<ts.length;j++){var t=ts[j],k=L.indexOf(t);if(k>=0){matched.push(t);var c=0,p=k;while(p>=0){c++;p=L.indexOf(t,p+t.length);}occ+=c;}}
  if(matched.length){var score=matched.length*1e6+Math.min(occ,999)*100+(r.w?r.w%10000:0)/10000;cur.push({r:r,matched:matched,score:score});}
 }
 cur.sort(function(a,b){return b.score-a.score;});
 render(ts,rx);
}
function snippet(r,rx){
 var x=r.x;rx.lastIndex=0;var m=rx.exec(r._l);var first=m?m.index:0;
 var start=Math.max(0,first-70),end=Math.min(x.length,first+210);
 var s=esc(x.slice(start,end)).replace(rx,'<mark>$1</mark>');
 return (start>0?'… ':'')+s+(end<x.length?' …':'');
}
function render(ts,rx){
 if(!ts||!ts.length){out.innerHTML='<p class="empty">키워드를 입력하면 결과가 여기에 표시됩니다.<br>총 <b>'+IDX.length+'</b>개 항목이 색인되어 있습니다. 한글·영문을 함께 적으면 더 잘 찾습니다.</p>';clr.style.display='none';return;}
 clr.style.display='block';
 var tchips=ts.map(function(t){var n=cur.filter(function(h){return h.matched.indexOf(t)>=0;}).length;return '<span class="chip">'+esc(t)+'<span class="n">'+n+'</span></span>';}).join(' ');
 var label='‘'+ts.map(esc).join('’ 또는 ‘')+'’';
 if(!cur.length){out.innerHTML='<p class="sumline">'+label+' — 결과 없음</p><div class="chips">'+tchips+'</div><p class="empty">일치하는 내용이 없습니다. 다른 키워드나 영문/한글 표기를 함께 입력해 보세요.</p>';return;}
 var cap=Math.min(cur.length,200),html='';
 for(var i=0;i<cap;i++){var h=cur[i],r=h.r,ext=r.e?' target="_blank" rel="noopener"':'';
  html+='<a class="res" href="'+r.u+'"'+ext+'><div class="top"><span class="badge">'+esc(r.g)+'</span><span class="ttl">'+esc(r.t)+'</span><span class="arrow">'+h.matched.length+'/'+ts.length+' 일치 →</span></div><div class="snip">'+snippet(r,rx)+'</div></a>';
 }
 if(cur.length>cap)html+='<p class="empty">상위 '+cap+'개만 표시했습니다. 키워드를 더해 좁혀 보세요.</p>';
 out.innerHTML='<p class="sumline">'+label+' — <b>'+cur.length+'</b>개 결과</p><div class="chips">'+tchips+'</div>'+html;
}
var deb;function onInput(){clearTimeout(deb);deb=setTimeout(function(){run(qEl.value);sync();},110);}
function sync(){try{history.replaceState(null,'',location.pathname+(qEl.value?'?q='+encodeURIComponent(qEl.value):''));}catch(e){}}
qEl.addEventListener('input',onInput);
clr.addEventListener('click',function(){qEl.value='';qEl.focus();run('');sync();});
[].forEach.call(document.querySelectorAll('.chip[data-q]'),function(c){c.addEventListener('click',function(){qEl.value=c.getAttribute('data-q');qEl.focus();run(qEl.value);sync();});});
qEl.addEventListener('keydown',function(e){
 var items=out.querySelectorAll('a.res');
 if(e.key==='ArrowDown'||e.key==='ArrowUp'){if(!items.length)return;e.preventDefault();sel+=e.key==='ArrowDown'?1:-1;if(sel<0)sel=items.length-1;if(sel>=items.length)sel=0;[].forEach.call(items,function(it){it.classList.remove('sel');});items[sel].classList.add('sel');items[sel].scrollIntoView({block:'nearest'});}
 else if(e.key==='Enter'){if(sel>=0&&items[sel])items[sel].click();else if(items[0])items[0].click();}
 else if(e.key==='Escape'){qEl.value='';run('');sync();}
});
var q0=new URLSearchParams(location.search).get('q')||'';qEl.value=q0;run(q0);
})();'''

_sugg = ['테슬라 TSLA','금리 yield','인플레이션 inflation','유동성 liquidity','로보택시 robotaxi','이더리움 ETH BMNR','반도체 SOXX','가이던스 guidance']
_chips = ''.join(f'<button class="chip" data-q="{esc(s)}">{esc(s)}</button>' for s in _sugg)
_search_top = ('<div class="card"><h2>🔍 통합 검색</h2>'
 '<p class="hint">한글·영문 키워드로 Summary 전체(주간 브리핑·Orientation·Factset·Yield·퀵 링크)를 검색합니다. '
 '여러 단어는 띄어쓰기로 구분하며 <b>OR</b>(하나라도 포함) 조건으로 찾고, 더 많은 단어가 일치할수록 위로 정렬됩니다.</p>'
 '<div class="sbox"><input id="q" type="text" placeholder="예) 테슬라 TSLA 로보택시" autocomplete="off" autofocus>'
 '<button class="clr" id="clr" title="지우기">×</button></div>'
 '<div class="chips">' + _chips + '</div>'
 '<p class="hint"><kbd>↑</kbd> <kbd>↓</kbd> 이동 · <kbd>Enter</kbd> 열기 · <kbd>Esc</kbd> 지우기</p></div>'
 '<div id="out"></div>')
_index_js = json.dumps(SEARCH_RECORDS, ensure_ascii=False).replace('</', '<\\/')
search_body = (SEARCH_CSS + _search_top
 + '<script>const SEARCH_INDEX=' + _index_js + ';</script>\n'
 + '<script>' + SEARCH_JS + '</script>')
open(os.path.join(SITE,'search.html'),'w',encoding='utf-8').write(page('검색','search', search_body))
print('search index records:', len(SEARCH_RECORDS))
print('site built OK (privacy filter active)')
